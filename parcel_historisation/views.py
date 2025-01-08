from django.conf import settings
from django.contrib.auth.decorators import permission_required
from django.forms.models import model_to_dict
from django.http import FileResponse, HttpResponse, JsonResponse
from django.template import loader
from rest_framework import viewsets, filters
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from openpyxl import load_workbook
from io import BytesIO

import json
import pathlib
import datetime

from parcel_historisation.models import Plan, Operation, OtherOperation, State, DivisonReunion, Balance
from parcel_historisation.models import VBalanceSourceNoDest, VBalanceDestNoSource
from parcel_historisation.serializers import PlanSerializer, OperationSerializer, BalanceSerializer


class PlanViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint that allows plans to be listed in read-only mode.
    """

    serializer_class = PlanSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["plan", "designation", "state", "date_plan"]

    ordering_fields = ["plan", "designation", "state", "date_plan"]
    ordering = ["date_plan"]

    def get_queryset(self):
        if self.request.query_params.get("numcad"):
            numcad = int(self.request.query_params.get("numcad"))
            queryset = Plan.objects.filter(cadastre=numcad).all()
        else:
            queryset = Plan.objects.all()

        return queryset


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def index(request):
    """
    Serving the base template.
    """

    template = loader.get_template("parcel_historisation/index.html")

    context = {"infolica_api_url": settings.INTRANET_PROXY.get("infolica_api_url")}

    return HttpResponse(template.render(context, request))


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def get_docs_list(request):
    """
    Gets the list of documents which have to be analysed, thus having a state
    equal to one. The list is generated regarding a specified cadastre
    """

    # TODO: parameter validation (try int(numcad) except BadRequest)
    numcad = request.GET["numcad"]

    results = Plan.objects.filter(state__id=1).filter(cadastre=int(numcad)).order_by("-date_plan", "link").all()

    load = []
    for result in results:
        load.append(
            {
                "link": result.link,
                "id": result.id,
                "designation_id": result.designation.id if result.designation else None,
            }
        )
    first = results[0]

    if first.designation is not None:
        download = first.designation.name
    else:
        download = first.name

    return JsonResponse({"list": load, "download": download}, safe=False)


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def file_download(request, name):
    """
    Data download view

    The whole path to the file has to be passed in the URL
    """

    base_path = settings.NEARCH2_CONSULTATION

    path = pathlib.Path(base_path + "/Plans_de_mutations/" + name)

    return FileResponse(open(path, "rb"))


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def get_download_path(request):
    """
    For a plan (or a designation -> depends on the type parameter),
    gets the whole download path (see file_download function)
    """

    id_ = request.GET["id"]
    type_ = request.GET["type"]

    object_ = Plan.objects.get(pk=id_)

    if type_ == "plan":
        download_path = object_.name
    elif type_ == "designation":
        download_path = object_.designation.name

    return JsonResponse({"download_path": download_path}, safe=False)


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def submit_saisie(request):
    """
    Process data submitted by user in the base form
    """

    data = json.loads(request.body)
    has_div = False
    plan = Plan.objects.get(pk=data["id"])

    if data["delayed_check"] is True:
        state = State.objects.get(pk=3)
        plan.state = state
        plan.save()
        return JsonResponse({"submitted": False, "has_div": has_div}, safe=False)

    now = datetime.datetime.now()
    op = Operation(date=now, user=request.META["HTTP_REMOTE_USER"], complement=data["complement"], plan=plan)

    operation_id = data["operation_id"] if "operation_id" in data else None
    if operation_id is not None:
        op.id = operation_id
    op.save()

    if data["div_check"] is True:
        div = DivisonReunion(operation=op)
        if not DivisonReunion.objects.filter(operation=div.operation).exists():
            div.save()
        has_div = True
    else:
        divs = DivisonReunion.objects.filter(operation=op).all()
        if len(divs) > 0:
            for div in divs:
                bal = Balance.objects.filter(division=div.id).all()  ## !! Cannot use QuerySet for "DivisonReunion": Use a QuerySet for "Operation"

                # remove balance if exists
                for tmp in bal:
                    tmp.delete()

                # remove division
                for tmp in div:
                    tmp.delete()

    if data["cad_check"] is True:
        other = OtherOperation(operation=op, type=1)
        if not OtherOperation.objects.filter(operation=other.operation, type=other.type).exists():
            other.save()
    else:
        op_type = OtherOperation.objects.filter(operation=op).all()
        for tmp in op_type:
            tmp.delete()

    if data["serv_check"] is True:
        other = OtherOperation(operation=op, type=2)
        if not OtherOperation.objects.filter(operation=other.operation, type=other.type).exists():
            other.save()
    else:
        op_type = OtherOperation.objects.filter(operation=op).all()
        for tmp in op_type:
            tmp.delete()

    if data["art35_check"] is True:
        other = OtherOperation(operation=op, type=3)
        if not OtherOperation.objects.filter(operation=other.operation, type=other.type).exists():
            other.save()
    else:
        op_type = OtherOperation.objects.filter(operation=op).all()
        for tmp in op_type:
            tmp.delete()

    if data["other_check"] is True:
        other = OtherOperation(operation=op, type=4)
        if not OtherOperation.objects.filter(operation=other.operation, type=other.type).exists():
            other.save()
    else:
        op_type = OtherOperation.objects.filter(operation=op).all()
        for tmp in op_type:
            tmp.delete()

    if data["cad_check"] is True:
        state = State.objects.get(pk=4)
    else:
        state = State.objects.get(pk=2)

    plan.state = state
    plan.save()

    return JsonResponse(
        {
            "submitted": True,
            "has_div": has_div,
            "operation_id": op.id,
        }
    )


def _build_balance(relations):
    source_bf = []
    destination_bf = []
    relations_ = []

    for rel in relations:
        if rel["destination"]:
            source_bf.append(rel["source"]) if not rel["source"] in source_bf else None
            destination_bf.append(rel["destination"]) if not rel["destination"] in destination_bf else None
            relations_.append([rel["source"], rel["destination"]])

    # rename "0_1" into "DP" and rename "0_2" into "RP"
    source_bf = ["DP" if x == "0_1" else x for x in source_bf]
    destination_bf = ["DP" if x == "0_1" else x for x in destination_bf]
    relations_ = [["DP" if y == "0_1" else y for y in x] for x in relations_]
    source_bf = ["RP" if x == "0_2" else x for x in source_bf]
    destination_bf = ["RP" if x == "0_2" else x for x in destination_bf]
    relations_ = [["RP" if y == "0_2" else y for y in x] for x in relations_]

    source_bf.sort()
    destination_bf.sort()

    balance = None

    if len(source_bf) > 0 and len(destination_bf) > 0:
        # initialize balance 2d list
        balance = [[" " for j in range(len(destination_bf) + 1)] for i in range(len(source_bf) + 1)]

        # 1rst line and row for headers
        for i in range(len(source_bf)):
            balance[i + 1][0] = source_bf[i]

        for j in range(len(destination_bf)):
            balance[0][j + 1] = destination_bf[j]

        for rel in relations_:
            src_idx = source_bf.index(rel[0]) + 1
            dst_idx = destination_bf.index(rel[1]) + 1
            balance[src_idx][dst_idx] = "X"

    return balance


class BalanceViewSet(viewsets.ViewSet):
    """
    API endpoint to handle balances.
    """

    serializer_class = BalanceSerializer
    queryset = Balance.objects.all()

    def retrieve(self, request, pk):
        instances = Balance.objects.filter(division=pk)
        relations = {}
        relations["balance"] = []
        relations["ddp"] = []
        for rel in instances:
            if rel.destination_ddp is True:
                relations["ddp"].append("-".join([rel.source, rel.destination]))
            else:
                relations["balance"].append({"source": rel.source, "destination": rel.destination})

        # Sort results
        relations["balance"] = _build_balance(relations["balance"])
        relations["ddp"] = sorted(
            relations["ddp"],
            key=lambda item: (
                int(item.split("-")[1].split("_")[0]),
                int(item.split("-")[1].split("_")[1]),
                int(item.split("_")[0]),
                int(item.split("_")[1].split("-")[0]),
            ),
        )

        return Response(relations)


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def submit_balance(request):
    """
    save balance
    """

    data = json.loads(request.body)

    # remove existing relations and (re-)create them
    data2remove = Balance.objects.filter(division_id=data["division_id"]).all()
    for _ in data2remove:
        _.delete()

    # save balance
    if data["balance"] is not None:
        for relation in data["balance"]:
            balance = Balance()
            rel = relation.split("-")

            # control if dp is in old/new parcel of relation
            for i, parcel in enumerate(rel):
                if "dp" in parcel.lower():
                    rel[i] = "DP"

            balance.source = rel[0]
            balance.destination = rel[1]
            balance.division_id = data["division_id"]

            balance.save()

    # save ddp relations
    if data["ddp"] is not None:
        for relation in data["ddp"]:
            balance = Balance()
            rel = relation.split("-")

            balance.source = rel[0]
            balance.destination = rel[1]
            balance.destination_ddp = True
            balance.division_id = data["division_id"]

            balance.save()

    return JsonResponse(
        {
            "submitted": True,
        }
    )


@permission_required("parcel_historisation.view_designation", raise_exception=True)
@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def balance_file_upload(request):
    """
    Upload balance file
    """
    cadnum = request.POST.get("cadnum")
    file = request.FILES.get("file")

    wb = load_workbook(filename=BytesIO(file.read()))
    ws = wb.active

    # get old parcels
    new_bf = []
    i = 1
    while ws.cell(1, i + 2).value is not None:
        new_bf.append(ws.cell(1, i + 2).value)
        i += 1
    new_bf = list(set(new_bf))
    if "dp" in new_bf:
        new_bf.remove("dp")
        new_bf.sort()
        new_bf.append("dp")
    else:
        new_bf.sort()

    # get old parcels
    old_bf = []
    i = 1
    while ws.cell(i + 1, 2).value is not None:
        old_bf.append(ws.cell(i + 1, 2).value)
        i += 1
    old_bf = list(set(old_bf))
    if "dp" in old_bf:
        old_bf.remove("dp")
        old_bf.sort()
        old_bf.append("dp")
    else:
        old_bf.sort()

    relations = []
    for row_id in range(len(old_bf)):
        for col_id in range(len(new_bf)):
            if ws.cell(row_id + 2, col_id + 3).value is not None:
                relations.append([ws.cell(row_id + 2, 2).value, ws.cell(1, col_id + 3).value])

    balance = None
    if len(old_bf) > 0 and len(new_bf) > 0:
        # initialize balance 2d list
        balance = [[" " for j in range(len(new_bf) + 1)] for i in range(len(old_bf) + 1)]

        # 1rst line and row for headers
        for i in range(len(old_bf)):
            tmp = str(old_bf[i])
            if not str(old_bf[i]) == "dp":
                tmp = "_".join([str(cadnum), str(old_bf[i])])
            balance[i + 1][0] = tmp

        for j in range(len(new_bf)):
            tmp = str(new_bf[j])
            if not str(new_bf[j]) == "dp":
                tmp = "_".join([str(cadnum), str(new_bf[j])])
            balance[0][j + 1] = tmp

        for rel in relations:
            src_idx = old_bf.index(rel[0]) + 1
            dst_idx = new_bf.index(rel[1]) + 1
            balance[src_idx][dst_idx] = "X"

    return JsonResponse({"balance": balance})


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def load_operation(request):
    """
    Load operation
    """
    operation_id = request.GET.get("operation")

    # check if operation_id already exists in database
    op = Operation.objects.get(id=operation_id)

    return JsonResponse(model_to_dict(op), safe=False)


class OperationViewSet(viewsets.ModelViewSet):
    """
    API endpoint that exposes operation in order to continue the edition mode
    """

    queryset = Operation.objects.all()
    serializer_class = OperationSerializer


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def liberate(request):
    """
    Liberate an operation (give again access to it)
    """
    id_ = request.GET.get("id")

    # check if operation_id already exists in database
    plan = Plan.objects.get(id=id_)
    plan.state = State(id=1)
    plan.save()

    return JsonResponse({"operation": 'saved'}, safe=False)


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def run_control(request, cad_no):
    """
    Returns everything which is a source but has no destination (1)
    AND everything which is a destination but has no source (2)

    (1):
    SELECT src.source
    FROM parcel_historisation.balance src
    LEFT JOIN parcel_historisation.balance dst
    ON src.source = dst.destination
    WHERE dst.destination is null
    AND src.source like '14_%'
    GROUP BY src.source

    (2):
    SELECT dst.destination
    FROM parcel_historisation.balance src
    -- RIGHT JOIN (SELECT * FROM parcel_historisation.balance WHERE destination like '14_%') dst
    RIGHT JOIN parcel_historisation.balance dst
    ON src.source = dst.destination
    WHERE src.source is null
    AND dst.destination like '14_%'
    GROUP BY dst.destination
    """

    src = VBalanceSourceNoDest.objects.filter(source__contains=cad_no+"_").all()
    vbalancesourcenodest = []
    for row in src:
        vbalancesourcenodest.append(row.pk)

    dst = VBalanceDestNoSource.objects.filter(destination__contains=cad_no+"_").all()
    vbalancedestnosource = []
    for row in dst:
        vbalancedestnosource.append(row.pk)

    return JsonResponse({
        "operation": cad_no,
        "sources": vbalancesourcenodest,
        "destination": vbalancedestnosource}, safe=False)


@permission_required("parcel_historisation.view_designation", raise_exception=True)
def submit_control(request):
    """
    Process control types submitted by user in the control panel
    """

    data = json.loads(request.body)

    is_ok = False

    if data["type"] in ('rp_out', 'origin'):
        for item in data["source_bfs"]:
            balance = Balance()
            balance.destination_ddp = False
            balance.destination = item
            balance.source = 'origin'
            if data["type"] == 'rp_out':
                balance.source_rp = True
            else:
                balance.source_origin = True
            balance.save()

        is_ok = True

    if data["type"] in ('rp_in', 'final'):
        for item in data["destination_bfs"]:
            balance = Balance()
            balance.destination_ddp = False
            balance.source = item
            balance.destination = 'destination'
            if data["type"] == 'rp_in':
                balance.destination_rp = True
            else:
                balance.current_destination = True
            balance.save()

        is_ok = True

    saved = False
    if is_ok is True:
        balance.save()
        saved = True

    return JsonResponse(
        {
            "submitted": True,
            "saved": saved,
        }
    )
